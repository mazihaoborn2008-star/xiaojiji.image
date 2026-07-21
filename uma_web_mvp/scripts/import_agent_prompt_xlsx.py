from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any


COLUMN_ALIASES = {
    "category": {"category", "分类", "类别", "类型"},
    "scene": {"scene", "场景"},
    "style": {"style", "风格", "画风"},
    "prompt": {"prompt", "提示词", "正向提示词", "正面提示词"},
    "negative": {"negative", "negative prompt", "负面词", "反向提示词", "负面提示词"},
    "tags": {"tags", "tag", "关键词", "标签"},
    "notes": {"notes", "备注", "说明"},
}


def normalize_header(value: Any) -> str:
    return str(value or "").strip().lower()


def detect_columns(headers: list[Any]) -> dict[str, int]:
    normalized = [normalize_header(value) for value in headers]
    result: dict[str, int] = {}
    for key, aliases in COLUMN_ALIASES.items():
        for index, header in enumerate(normalized):
            if header in aliases:
                result[key] = index
                break
    return result


def import_xlsx(source: Path, output: Path) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise SystemExit("openpyxl is required. Install dependencies from requirements.txt first.") from exc

    workbook = load_workbook(source, read_only=True, data_only=True)
    items: list[dict[str, str]] = []
    skipped_sheets = []
    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration:
            continue
        columns = detect_columns(list(headers))
        if "prompt" in columns or "tags" in columns:
            for row_index, row in enumerate(rows, start=2):
                item: dict[str, str] = {"id": f"{sheet.title}:{row_index}"}
                has_content = False
                for key in COLUMN_ALIASES:
                    column_index = columns.get(key)
                    value = "" if column_index is None or column_index >= len(row) else str(row[column_index] or "").strip()
                    if value:
                        has_content = True
                    item[key] = value
                if has_content:
                    items.append(item)
            continue

        pair_items = parse_tag_pairs(sheet.title, [headers], rows)
        if pair_items:
            items.extend(pair_items)
        else:
            skipped_sheets.append(sheet.title)

    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(
        json.dumps({"source": str(source), "count": len(items), "items": items}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return {"count": len(items), "output": str(output), "skipped_sheets": skipped_sheets}


def parse_tag_pairs(sheet_title: str, first_rows: list[tuple[Any, ...]], rows_iter) -> list[dict[str, str]]:
    if "和谐" in sheet_title:
        return []
    result: list[dict[str, str]] = []
    rows = list(first_rows)
    current_sections: dict[int, str] = {}
    for index, row in enumerate(rows_iter, start=2):
        rows.append(row)
        if len(rows) > 20000:
            break
    for row_index, row in enumerate(rows, start=1):
        values = [str(cell or "").strip() for cell in row]
        for col, value in enumerate(values):
            if not value:
                continue
            if not is_probable_tag(value):
                if value and len(value) <= 30 and not any(ch.isascii() and ch.isalpha() for ch in value):
                    current_sections[col] = value
                continue
            zh = ""
            if col + 1 < len(values) and values[col + 1] and not is_probable_tag(values[col + 1]):
                zh = values[col + 1]
            section = current_sections.get(col) or current_sections.get(max(0, col - 1)) or ""
            result.append(
                {
                    "id": f"{sheet_title}:{row_index}:{col + 1}",
                    "category": sheet_title,
                    "scene": section,
                    "style": "",
                    "prompt": value,
                    "negative": "",
                    "tags": value,
                    "notes": zh,
                }
            )
    return result


def is_probable_tag(value: str) -> bool:
    value = value.strip()
    if len(value) > 120:
        return False
    if not re.search(r"[A-Za-z_]", value):
        return False
    if re.search(r"[\u4e00-\u9fff]", value):
        return False
    return bool(re.fullmatch(r"[A-Za-z0-9_\\/ (),.'’-]+", value))


def main() -> None:
    parser = argparse.ArgumentParser(description="Import Smart Agent prompt snippets from agent.xlsx.")
    parser.add_argument("--source", default=r"C:\Users\Administrator\Desktop\agent.xlsx")
    parser.add_argument("--output", default=str(Path(__file__).resolve().parents[1] / "app" / "data" / "agent_prompt_library.json"))
    args = parser.parse_args()
    result = import_xlsx(Path(args.source), Path(args.output))
    print(
        "imported={count} output={output} skipped_sheets={skipped}".format(
            count=result["count"],
            output=result["output"],
            skipped=len(result["skipped_sheets"]),
        )
    )


if __name__ == "__main__":
    main()
